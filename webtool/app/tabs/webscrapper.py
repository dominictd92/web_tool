#
# This is a webscrapper that can cycle through any
#   list of urls and look for a specific string within those urls.
#

import os
import threading
import requests
from tkinter import (W, EW, NSEW, LEFT, ALL, INSERT, END, NORMAL, DISABLED, IntVar, BooleanVar, StringVar, Button,
                     Radiobutton, Checkbutton, Entry, Label, LabelFrame, Frame)
from tkinter.messagebox import showinfo
from tkinter.filedialog import askopenfilename

from openpyxl import load_workbook

from webtool.constants.constants import default_history_filename, default_results_filename, excel_ext
from webtool.widget.CollapsibleFrame import CollapsibleFrame
from webtool.widget.Hint import Hint
from webtool.widget.ScrollableFrame import ScrollableFrame
from webtool.widget.ScrollableText import ScrollableText
from webtool.utils.fileutils import get_default_search_file, export_txt, export_xlsx, upload_file
from webtool.utils.loggingutils import get_logger
from webtool.utils.validationutils import validate_int
from webtool.utils.webutils import return_webpage


# the class for the webscrapper with all needed methods within it
class Webscrapper(Frame):
    # class variables
    found_text_formatter = " - Found Text: '{search_string}' on {url} {find_count} time(s)\n"
    history_text_formatter = "Searches: {searches}\n{results}" + ("-" * 15) + "\n"

    notes = (
        "- The webscrapper can traverse through any list of urls and search for any text within the HTML.\n"
        "- The file has to be in an excel format and listed in the first column.\n"
        "- The default file to be searched is 'search_file.xlsx' in the program's directory.\n"
        "- To search for multiple things, use '+' with no spaces between the '+' and the search text.\n"
        "- Hover over items with \u2071 alongside it.")

    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        super().columnconfigure(0, weight=1)
        super().rowconfigure(0, weight=1)

        self.logger = get_logger()

        self.scrollable_frame = ScrollableFrame(self, *args, **kwargs)

        # Variables used throughout the webscrapper
        self.found_results = []
        self.filename = StringVar()
        self.filename.set(get_default_search_file())
        self.input_variable = StringVar()
        self.completion_text = StringVar()
        self.completion_text.set("Results: ")
        note_text = StringVar()
        note_text.set(self.notes)
        self.stop_variable = BooleanVar()
        self.stop_variable.set(False)
        # variables for options
        self.specific_tags = StringVar()
        self.specific_tags.set("Look for specific tags (in development...)")
        self.case_sensitive = IntVar()
        self.retry_urls = IntVar()
        self.retry_urls_var = IntVar()
        self.default_timeout = IntVar()

        # Webscrapper frames
        #   Note frame
        note_frame = Frame(self.scrollable_frame, highlightbackground='white', highlightthickness=1)
        note_frame.grid(row=0, column=0, sticky=EW)

        #   Search frame
        search_frame = Frame(self.scrollable_frame, highlightbackground='white', highlightthickness=1)
        search_frame.grid(row=1, column=0, sticky=EW)

        #   Options frame
        options_frame = Frame(self.scrollable_frame, highlightbackground='white', highlightthickness=1)
        options_frame.grid(row=2, column=0, sticky=EW)

        #   Buttons frame
        button_frame = Frame(self.scrollable_frame)
        button_frame.grid(row=3, column=0, sticky=EW)
        button_frame.columnconfigure(2, weight=1)

        #   Results frame
        self.results_labelframe = LabelFrame(
            self.scrollable_frame, text=self.completion_text.get(), height=200, width=725)
        self.results_labelframe.grid(row=4, column=0, sticky=EW)
        self.results_labelframe.grid_rowconfigure(0, weight=1)
        self.results_labelframe.grid_columnconfigure(0, weight=1)
        self.results_labelframe.grid_propagate(False)

        #   Skipped urls frame
        skipped_labelframe = LabelFrame(
            self.scrollable_frame, text="Skipped URLs:", height=200, width=725)
        skipped_labelframe.grid(row=5, column=0, sticky=EW)
        skipped_labelframe.grid_rowconfigure(0, weight=1)
        skipped_labelframe.grid_columnconfigure(0, weight=1)
        skipped_labelframe.grid_propagate(False)

        #   History frame
        history_labelframe = LabelFrame(
            self.scrollable_frame, text="History:", height=200, width=725)
        history_labelframe.grid(row=6, column=0, sticky=EW)
        history_labelframe.grid_rowconfigure(0, weight=1)
        history_labelframe.grid_columnconfigure(0, weight=1)
        history_labelframe.grid_propagate(False)

        # Widgets within the frames
        #   Note frame
        note_label_label = Label(
            note_frame, text="", justify=LEFT, cursor='hand2')
        note_label_label.grid(row=0, column=0, padx=5, pady=5, sticky=W)

        note_collapsible_frame = CollapsibleFrame(note_frame, "Note", note_label_label)
        note_collapsible_frame.grid(row=1, column=0, padx=5, pady=5, sticky=EW)

        note_label = Label(note_collapsible_frame, textvariable=note_text, wraplength=680, justify=LEFT)
        note_label.grid(row=0, column=0, columnspan=5, padx=10, pady=(0, 10))

        #   Search frame
        search_label = Label(search_frame, text="Search Settings")
        search_label.grid(row=0, column=0, sticky=W, padx=(5, 0), pady=5)

        select_file_label = Label(search_frame, text="URL File:", justify=LEFT, anchor=W)
        select_file_label.grid(row=1, column=0, padx=(15, 5), pady=5, sticky=W)

        file_label = Label(search_frame, textvariable=self.filename, justify=LEFT, anchor=W)
        file_label.grid(row=1, column=1, columnspan=4, padx=5, pady=5, sticky=W)

        file_button = Button(
            search_frame, text="Browse Files", width=10,
            command=lambda: self.filename.set(
                    askopenfilename(title="Select a file", initialdir=os.getcwd(), filetypes=excel_ext)))
        file_button.grid(row=2, column=0, padx=(15, 5), pady=5, sticky=W)

        clear_file_button = Button(search_frame, text="Clear File", width=10,
                                   command=lambda: self.filename.set(get_default_search_file()))
        clear_file_button.grid(row=2, column=1, padx=5, pady=5, sticky=W)

        upload_file_button = Button(search_frame, text="Upload File", width=10,
                                    command=lambda: upload_file(extensions=excel_ext))
        upload_file_button.grid(row=2, column=2, padx=5, pady=5, sticky=W)

        # needed for the column size/position of upload button
        search_frame.columnconfigure(2, weight=1)

        search_label = Label(search_frame, text="Search terms: ", justify=LEFT)
        search_label.grid(row=3, column=0, padx=(15, 5), pady=5, sticky=W)

        search_entry = Entry(search_frame, textvariable=self.input_variable, width=50, justify=LEFT)
        search_entry.grid(row=3, column=1, columnspan=4, padx=5, pady=5, sticky=W)

        #   Options frame
        options_label_label = Label(options_frame, text="", anchor=W, justify=LEFT, cursor='hand2')
        options_label_label.grid(row=0, column=0, padx=5, pady=5, sticky=W)

        options_collapsible_frame = CollapsibleFrame(options_frame, "Options", options_label_label)
        options_collapsible_frame.grid(row=1, sticky=EW)

        specific_tags_label = Label(options_collapsible_frame, text="HTML tags \u2071")
        specific_tags_label.grid(row=0, column=0, sticky=W, padx=(15, 5), pady=5)

        Hint(specific_tags_label, "In development...")

        specific_tags_entry = Entry(options_collapsible_frame, textvariable=self.specific_tags,
                                    width=50, justify=LEFT, state=DISABLED)
        specific_tags_entry.grid(row=0, column=1, columnspan=4, sticky=W, padx=5, pady=5)

        option_case_sensitive = Checkbutton(options_collapsible_frame, text="Case Sensitive",
                                            variable=self.case_sensitive, onvalue=1, offvalue=0)
        option_case_sensitive.grid(row=1, column=0, sticky=W, padx=(15, 5), pady=5)

        self.option_retry_urls = Checkbutton(
            options_collapsible_frame, text="Retry Skipped Urls?\u2071",
            variable=self.retry_urls, onvalue=1, offvalue=0, command=lambda: self.toggle_retry_radio_buttons())
        self.option_retry_urls.grid(row=2, column=0, sticky=W, padx=(15, 5), pady=5)

        Hint(self.option_retry_urls, "Retry urls that are skipped due to timeout?")

        self.option_auto_retry_urls = Radiobutton(
            options_collapsible_frame, state=DISABLED, text="Retry after process\u2071",
            variable=self.retry_urls_var, value=1)
        self.option_auto_retry_urls.grid(row=2, column=1, sticky=W, padx=10, pady=5)

        Hint(self.option_auto_retry_urls, "Retry skipped urls after running through the entire file")

        self.option_force_retry_urls = Radiobutton(
            options_collapsible_frame, state=DISABLED, text="Retry now\u2071",
            variable=self.retry_urls_var, value=2)
        self.option_force_retry_urls.grid(row=2, column=2, sticky=W, padx=10, pady=5)

        Hint(self.option_force_retry_urls, "Retry skipped urls now, ignoring the Excel file")

        option_default_timeout = Checkbutton(
            options_collapsible_frame, text="Default Timeout", variable=self.default_timeout,
            onvalue=1, offvalue=0, command=lambda: self.toggle_default_retries())
        option_default_timeout.grid(row=3, column=0, sticky=W, padx=(15, 5), pady=5)

        option_retry_seconds_label = Label(options_collapsible_frame, text="Seconds before timeout \u2071")
        option_retry_seconds_label.grid(row=3, column=1, sticky=W, padx=10, pady=5)

        Hint(option_retry_seconds_label, "Default: 3, Min: 2, Max: 6")

        reg_option_retry_seconds_entry = (options_collapsible_frame.register(validate_int),
                                          '%P', '%V', '2', '6')

        self.option_retry_seconds_entry = Entry(
            options_collapsible_frame, width=3, justify=LEFT,
            validate=ALL, validatecommand=reg_option_retry_seconds_entry)
        self.option_retry_seconds_entry.grid(row=3, column=2, sticky=W, padx=5, pady=5)
        self.option_retry_seconds_entry.insert(0, "3")

        option_retry_quantity_label = Label(options_collapsible_frame, text="Initial timeout retries \u2071")
        option_retry_quantity_label.grid(row=3, column=3, sticky=W, padx=5, pady=5)

        Hint(option_retry_quantity_label, "Default: 3, Min: 0, Max: 5")

        reg_option_retry_quantity_entry = (options_collapsible_frame.register(validate_int),
                                           '%P', '%V', '0', '5')

        self.option_retry_quantity_entry = Entry(
            options_collapsible_frame, width=3, justify=LEFT,
            validate=ALL, validatecommand=reg_option_retry_quantity_entry)
        self.option_retry_quantity_entry.grid(row=3, column=4, columnspan=4, sticky=W, padx=5, pady=5)
        self.option_retry_quantity_entry.insert(0, "3")

        #   Buttons frame
        self.search_button = Button(
            button_frame, text="Search", state=NORMAL, width=10, fg="black", disabledforeground="black",
            command=lambda: False if not self.validate_values() else threading.Thread(target=self.scrape).start())
        self.search_button.grid(row=0, column=0, padx=10, pady=5)

        self.stop_button = Button(
            button_frame, text="Stop", state=DISABLED, width=10, fg="black", disabledforeground="black",
            command=lambda: self.stop_variable.set(True))
        self.stop_button.grid(row=0, column=1, padx=10, pady=5)

        self.export_results_button = Button(
            button_frame, text="Export Results", state=NORMAL, width=10, fg="black", disabledforeground="black",
            command=lambda: threading.Thread(
                target=export_xlsx(
                    dataset=self.found_results,
                    default_filename=default_results_filename)).start())
        self.export_results_button.grid(row=0, column=3, padx=10, pady=5)

        export_history_button = Button(
            button_frame, text="Export History", width=10, fg="black", disabledforeground="black",
            command=lambda: threading.Thread(
                    target=export_txt(
                        dataset=self.history_text.get("1.0", END),
                        default_filename=default_history_filename)).start())
        export_history_button.grid(row=0, column=4, padx=10, pady=5)

        #   Results frame
        self.results_text = ScrollableText(self.results_labelframe, highlightthickness=0, wrap='none')
        self.results_text.grid(row=0, column=0, columnspan=4, sticky=NSEW, pady=(5, 5), padx=(5, 15))

        #   Skipped frame
        self.skipped_text = ScrollableText(skipped_labelframe, highlightthickness=0, wrap='none')
        self.skipped_text.grid(row=0, column=0, columnspan=4, sticky=NSEW, pady=(5, 5), padx=(5, 15))

        #   History frame
        self.history_text = ScrollableText(history_labelframe, highlightthickness=0, wrap='none')
        self.history_text.grid(row=0, column=0, columnspan=4, sticky=NSEW, pady=(5, 5), padx=(5, 15))

        # start with these collapsible frames closed
        note_collapsible_frame.activate()
        options_collapsible_frame.activate()

    # Change the state of the search button, export results button, and stop button
    def flip_buttons(self):
        if str(self.search_button['state']) == str(DISABLED):
            self.search_button['state'] = str(NORMAL)
            self.export_results_button['state'] = str(NORMAL)
            self.stop_button['state'] = str(DISABLED)
            self.stop_variable.set(False)
        else:
            self.search_button['state'] = str(DISABLED)
            self.export_results_button['state'] = str(DISABLED)
            self.stop_button['state'] = str(NORMAL)

    # Toggle state of retry options based on default selected
    def toggle_default_retries(self):
        if self.default_timeout.get() == 0:
            self.option_retry_seconds_entry.config(state=NORMAL)
            self.option_retry_quantity_entry.config(state=NORMAL)
        else:
            self.option_retry_seconds_entry.delete(0, END)
            self.option_retry_seconds_entry.insert(0, "3")
            self.option_retry_seconds_entry.config(state=DISABLED)

            self.option_retry_quantity_entry.delete(0, END)
            self.option_retry_quantity_entry.insert(0, "3")
            self.option_retry_quantity_entry.config(state=DISABLED)

    def toggle_retry_radio_buttons(self):
        if self.retry_urls.get() == 1:
            self.retry_urls_var.set(1)
            self.option_auto_retry_urls.select()
            self.option_auto_retry_urls.config(state=NORMAL)
            self.option_force_retry_urls.config(state=NORMAL)
        else:
            self.retry_urls_var.set(0)
            self.option_auto_retry_urls.deselect()
            self.option_auto_retry_urls.config(state=DISABLED)
            self.option_force_retry_urls.deselect()
            self.option_force_retry_urls.config(state=DISABLED)

    # Validate all user inputs, returning true/false
    def validate_values(self):
        errors = []
        if not os.path.isfile(self.filename.get()):
            errors.append(f"- The file {self.filename.get()} could not be found.")

        if self.input_variable.get() == "":
            errors.append("- No search arguments entered.")

        if self.retry_urls_var.get() == 2 and len(self.skipped_text.get('1.0', END)) == 0:
            errors.append("- No skipped URLs to search.")

        if not validate_int(self.option_retry_quantity_entry.get()):
            errors.append("- The number of retries option is invalid")

        if not validate_int(self.option_retry_seconds_entry.get()):
            errors.append("- The seconds before timeout option is invalid")

        if len(errors) != 0:
            showinfo("Error", '\n'.join(errors))
            return False
        return True

    # function to search a website for
    def search_url(self, url, session, retries, wait_time):
        if url is not None and (url.find("https:") == 0 or url.find("http:") == 0):
            try:
                current_page = return_webpage(session, url, retry_quantity=retries, timeout_limit=wait_time)
                for search_string in self.input_variable.get().split("+"):
                    if self.case_sensitive.get() == 1:
                        find_count = current_page.text.count(search_string)
                    else:
                        find_count = current_page.text.upper().count(search_string.upper())

                    if find_count != 0:
                        found_text = self.found_text_formatter.format(
                            search_string=search_string,
                            url=url,
                            find_count=find_count)
                        self.found_results.append(found_text)
                        self.results_text.insert(END, found_text)
            except requests.exceptions.Timeout:
                self.logger.debug(f"Url request timeout: {url}")
                self.skipped_text.insert(INSERT, url + "\n")

    # scrape the web based on the user's selection
    def scrape(self):
        self.flip_buttons()
        # Reset variables for the new search
        http = requests.Session()  # Create a session object to use so that it is faster to reach
        self.found_results = []
        percentage_complete = 0
        percent_complete_text = format(percentage_complete, ".0%")
        self.results_text.delete('1.0', END)

        try:
            # only run the skipped urls if selected
            if self.retry_urls_var != 1:
                # get the file that has been selected
                url_file = load_workbook(self.filename.get(), read_only=True)
                worksheet = url_file.active

                for row in worksheet.iter_rows(1, worksheet.max_row):
                    if self.stop_variable.get():
                        self.logger.info("User stopped the webscrapper at " + percent_complete_text)
                        self.completion_text.set("Results: User stopped at " + percent_complete_text)
                        self.results_labelframe.configure(text=self.completion_text.get())
                        break
                    if round(row[0].row / worksheet.max_row, 2) > percentage_complete:
                        percentage_complete = round(row[0].row / worksheet.max_row, 2)
                        percentage_complete_text = format(percentage_complete, ".0%")
                        self.completion_text.set("Results: " + percentage_complete_text)
                        self.results_labelframe.configure(text=self.completion_text.get())
                    url = row[0].value

                    self.search_url(url, http, int(self.option_retry_quantity_entry.get()),
                                    int(self.option_retry_seconds_entry.get()))
                url_file.close()

            if self.retry_urls_var == 1 or self.retry_urls_var == 2:
                # run the skipped urls if the user chose to do so in any way
                skipped_urls = self.skipped_text.get('1.0', END).split("\n")
                self.skipped_text.delete('1.0', END)

                for index, skipped_url in enumerate(skipped_urls):
                    if self.stop_variable.get():
                        self.completion_text.set("Results: User stopped at " + percent_complete_text)
                        self.results_labelframe.configure(text=self.completion_text.get())
                        break
                    if round(index / len(skipped_urls), 2) > percentage_complete:
                        percentage_complete = round(index / len(skipped_urls), 2)
                        percentage_complete_text = format(percentage_complete, ".0%")
                        self.completion_text.set("Results: " + percentage_complete_text)
                        self.results_labelframe.configure(text=self.completion_text.get())

                    self.search_url(skipped_url, http, int(self.option_retry_quantity_entry.get()),
                                    int(self.option_retry_seconds_entry.get()))
            if len(self.found_results) == 0:
                message = "- The search produced no results (check skipped URLs)."
                self.results_text.delete('1.0', END)
                self.results_text.insert(INSERT, message)
                history_text = self.history_text_formatter.format(
                    searches=",".join(self.input_variable.get().split("+")),
                    results=message)
            else:
                history_text = self.history_text_formatter.format(
                    searches=", ".join(self.input_variable.get().split("+")),
                    results="".join(self.found_results))
            self.history_text.insert('1.0', history_text)

            http.close()
        except Exception as thrownException:
            self.logger.error(f"Unexpected error when scraping: {repr(thrownException)}")
        finally:
            self.flip_buttons()
